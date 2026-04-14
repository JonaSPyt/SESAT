"""
Módulo de importação de planilha Excel (.xlsx) para o banco de dados SESAT.
Lê todas as abas de um arquivo .xlsx no formato da planilha OPR e insere
os registros na tabela equipamentos, evitando duplicatas.
"""
from openpyxl import load_workbook
from datetime import datetime, timedelta, date
import re
import database


# Mapeamento flexível de cabeçalhos da planilha -> campos do banco
_HEADER_MAP = {
    # data_entrada
    "DATA DA ENTRADA":    "data_entrada",
    "DATA DA ENTREGA":    "data_entrada",
    "DATA ENTRADA":       "data_entrada",
    # secao_zona
    "SEÇÃO/ZONA":         "secao_zona",
    "SEÇÃO /ZONA":        "secao_zona",
    "SEÇÃO/ ZONA":        "secao_zona",
    "SECAO/ZONA":         "secao_zona",
    # tombamento
    "TOMBAMENTO":         "tombamento",
    # tipo
    "TIPO":               "tipo",
    # num_chamado
    "Nº CHAMADO":         "num_chamado",
    "N° CHAMADO":         "num_chamado",
    "NUM CHAMADO":        "num_chamado",
    "SOL":                "num_chamado",
    "OTRS":               "num_chamado",
    # tecnico_opr_entrada
    "TECNICO OPR ENTRADA": "tecnico_opr_entrada",
    "TÉCNICO OPR ENTRADA": "tecnico_opr_entrada",
    "TECNICO ENTRADA":     "tecnico_opr_entrada",
    "TÉCNICO- SEQUI":      "tecnico_opr_entrada",
    "TÉCNICO-SEQUI":       "tecnico_opr_entrada",
    "TÉCNICO SEQUI":       "tecnico_opr_entrada",
    # tecnico_sesat
    "TECNICO SESAT":       "tecnico_sesat",
    "TÉCNICO SESAT":       "tecnico_sesat",
    "TÉCNICO -SECAT":      "tecnico_sesat",
    "TÉCNICO-SECAT":       "tecnico_sesat",
    "TÉCNICO SECAT":       "tecnico_sesat",
    # tecnico_opr_devolucao
    "TECNICO OPR DEVOLUÇÃO":  "tecnico_opr_devolucao",
    "TÉCNICO OPR DEVOLUÇÃO":  "tecnico_opr_devolucao",
    "TECNICO DEVOLUÇÃO":      "tecnico_opr_devolucao",
    "TÉCNICO DEVOLUÇÃO":      "tecnico_opr_devolucao",
    "TECNICO DEVOLUCAO":      "tecnico_opr_devolucao",
    "TÉCNICO OPR DEVOLUCAO":  "tecnico_opr_devolucao",
    # data_saida
    "DATA DA SAÍDA":       "data_saida",
    "DATA DE SAÍDA":       "data_saida",
    "DATA SAÍDA":          "data_saida",
    "DATA DE SAIDA":       "data_saida",
    "DATA DA SAIDA":       "data_saida",
    "DATA SAIDA":          "data_saida",
    # observacoes
    "OBSERVAÇÕES":         "observacoes",
    "OBSERVACOES":         "observacoes",
    # anotacoes
    "ANOTAÇÕES DIVERSAS":  "anotacoes",
    "ANOTACOES DIVERSAS":  "anotacoes",
    "ANOTAÇÕES":           "anotacoes",
}

# Pré-computar mapa normalizado para evitar reprocessamento
_NORM_HEADER_MAP = {}


def _get_norm_header_map():
    """Retorna mapa normalizado (calculado uma vez)."""
    if not _NORM_HEADER_MAP:
        for header_texto, campo in _HEADER_MAP.items():
            norm = _normalizar_header(header_texto)
            _NORM_HEADER_MAP[norm] = campo
    return _NORM_HEADER_MAP


def _normalizar_header(texto: str) -> str:
    """Remove espaços extras e converte para maiúsculas para comparação."""
    return " ".join(texto.strip().upper().split())


def _excel_serial_to_date(serial) -> str:
    """Converte número serial do Excel para data no formato DD/MM/YY."""
    try:
        serial = float(serial)
        if serial < 1 or serial > 2958465:  # limites razoáveis (1900-9999)
            return ""
        # Excel epoch: 1900-01-01 (serial 1), com bug do 29/02/1900
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return dt.strftime("%d/%m/%y")
    except (ValueError, TypeError, OverflowError):
        return ""


def _parse_date_string(s: str) -> str:
    """Tenta interpretar uma string como data e retornar DD/MM/YY."""
    s = s.strip()
    if not s:
        return ""

    # Já no formato DD/MM/YY ou DD/MM/YYYY
    m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$', s)
    if m:
        dia, mes, ano = m.group(1), m.group(2), m.group(3)
        if len(ano) == 4:
            ano = ano[2:]  # 2019 -> 19
        return f"{int(dia):02d}/{int(mes):02d}/{ano}"

    # Formato YYYY-MM-DD (ISO)
    m = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$', s)
    if m:
        ano, mes, dia = m.group(1), m.group(2), m.group(3)
        return f"{int(dia):02d}/{int(mes):02d}/{ano[2:]}"

    return s


def _converter_valor_celula(valor, campo: str) -> str:
    """Converte o valor da célula para string, tratando datas e números."""
    if valor is None:
        return ""

    # Campos de data — podem vir como datetime, date, número serial ou string
    if campo in ("data_entrada", "data_saida"):
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%y")
        if isinstance(valor, date):
            return valor.strftime("%d/%m/%y")
        if isinstance(valor, (int, float)):
            return _excel_serial_to_date(valor)
        # Tentar interpretar string como data
        return _parse_date_string(str(valor))

    # Tombamento — manter como inteiro se vier como float
    if campo == "tombamento":
        if isinstance(valor, float) and valor == int(valor):
            return str(int(valor))
        return str(valor).strip()

    # Num chamado — pode ser número
    if campo == "num_chamado":
        if isinstance(valor, float) and valor == int(valor):
            return str(int(valor))
        return str(valor).strip()

    return str(valor).strip()


def importar_xlsx(caminho: str, usuario: str = "Supervisor") -> dict:
    """
    Importa registros de um arquivo .xlsx para o banco de dados.

    Retorna dict com:
        - total_abas: quantidade de abas processadas
        - total_importados: registros inseridos
        - total_duplicados: registros ignorados (já existiam)
        - total_ignorados: linhas vazias ou sem tombamento
        - erros: lista de strings descrevendo erros
    """
    wb = load_workbook(caminho, read_only=True, data_only=True)

    resultado = {
        "total_abas": 0,
        "total_importados": 0,
        "total_duplicados": 0,
        "total_ignorados": 0,
        "erros": [],
    }

    # Buscar chaves já existentes com query leve (só tombamento + data)
    existentes = database.buscar_chaves_existentes()

    norm_map = _get_norm_header_map()

    # Acumular registros para inserção em lote
    batch = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        resultado["total_abas"] += 1

        # Detectar linha de cabeçalho (procura nas 5 primeiras linhas)
        col_map = {}  # índice da coluna -> campo do banco
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=False), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    norm = _normalizar_header(cell.value)
                    campo = norm_map.get(norm)
                    if campo:
                        col_map[cell.column - 1] = campo
            if "tombamento" in col_map.values():
                header_row = row_idx
                break

        if header_row is None:
            resultado["erros"].append(
                f"Aba '{nome_aba}': cabeçalho não reconhecido, ignorada.")
            continue

        # Ler dados a partir da linha após o cabeçalho
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Extrair valores mapeados
            dados = {
                "data_entrada": "",
                "secao_zona": "",
                "nome_setor": "",
                "local_setor": "",
                "tombamento": "",
                "tipo": "",
                "descricao_completa": "",
                "nome_responsavel": "",
                "valor_unitario": "",
                "num_chamado": "",
                "tecnico_opr_entrada": "",
                "tecnico_sesat": "",
                "tecnico_opr_devolucao": "",
                "data_saida": "",
                "observacoes": "",
                "anotacoes": "",
            }

            for col_idx, campo in col_map.items():
                if col_idx < len(row):
                    dados[campo] = _converter_valor_celula(row[col_idx], campo)

            # Ignorar linhas sem tombamento
            if not dados["tombamento"]:
                resultado["total_ignorados"] += 1
                continue

            # Verificar duplicata (mesmo tombamento + data_entrada)
            chave = (dados["tombamento"], dados["data_entrada"])
            if chave in existentes:
                resultado["total_duplicados"] += 1
                continue

            # Adicionar ao lote
            batch.append(dados)
            existentes.add(chave)
            resultado["total_importados"] += 1

    wb.close()

    # Inserir todos de uma vez em uma única transação
    if batch:
        try:
            database.inserir_equipamentos_batch(batch)
        except Exception as e:
            resultado["erros"].append(f"Erro na inserção em lote: {e}")
            resultado["total_importados"] = 0

    # Registrar log da importação
    database.registrar_log(
        usuario=usuario,
        acao="IMPORTACAO",
        detalhes=(
            f"Importação XLSX: {resultado['total_importados']} inseridos, "
            f"{resultado['total_duplicados']} duplicados, "
            f"{resultado['total_ignorados']} ignorados, "
            f"{len(resultado['erros'])} erros"
        ),
    )

    return resultado
