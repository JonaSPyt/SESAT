#!/usr/bin/env python3
"""
SESAT - Sistema de Entrada e Saída de Equipamentos
Interface gráfica principal (Tkinter) — v2
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import calendar as cal_mod
import threading
import atexit

import database
import consulta_api
import exportar
import importar


# ═══════════════════════════════════════════════════════════════════
#  Paleta de cores (Catppuccin Mocha)
# ═══════════════════════════════════════════════════════════════════
BG = "#1e1e2e"
BG_SURFACE = "#24243a"
BG_CARD = "#2a2a3d"
BG_ENTRY = "#313244"
BG_HOVER = "#3b3b52"
BORDER = "#45475a"
FG = "#cdd6f4"
FG_DIM = "#a6adc8"
FG_MUTED = "#6c7086"
ACCENT = "#89b4fa"
ACCENT2 = "#74c7ec"
GREEN = "#a6e3a1"
RED = "#f38ba8"
YELLOW = "#f9e2af"
PEACH = "#fab387"
MAUVE = "#cba6f7"
HEADER_BG = "#181825"
ROW_EVEN = "#252538"
ROW_ODD = "#2e2e44"
SEL_BG = "#45475a"

FONT_FAMILY = "Segoe UI"
FONT = (FONT_FAMILY, 10)
FONT_BOLD = (FONT_FAMILY, 10, "bold")
FONT_TITLE = (FONT_FAMILY, 18, "bold")
FONT_SUB = (FONT_FAMILY, 11, "bold")
FONT_SMALL = (FONT_FAMILY, 9)
FONT_TINY = (FONT_FAMILY, 8)

ENTRY_OPTS = dict(
    font=FONT, bg=BG_ENTRY, fg=FG, insertbackground=FG,
    relief="flat", bd=0, highlightthickness=2,
    highlightbackground="#6272a4", highlightcolor=ACCENT,
)
LABEL_OPTS = dict(font=FONT_SMALL, bg=BG_CARD, fg=FG_DIM, anchor="w")
PADX = 12
PADY_FIELD = (6, 1)


class LoginWindow(tk.Tk):
    """Janela de login com usuário e senha."""

    def __init__(self):
        super().__init__()
        self.title("SESAT — Login")
        self.geometry("420x420")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.usuario = None
        self.is_super = False
        self.is_viewer = False

        # Centralizar na tela
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 210
        y = (self.winfo_screenheight() // 2) - 210
        self.geometry(f"420x420+{x}+{y}")

        self._build()

    def _build(self):
        # Logo / título
        tk.Label(
            self, text="⚙", font=(FONT_FAMILY, 36), bg=BG, fg=ACCENT
        ).pack(pady=(25, 5))
        tk.Label(
            self, text="SESAT", font=FONT_TITLE, bg=BG, fg=FG
        ).pack()
        tk.Label(
            self, text="Sistema de Entrada e Saída de Ativos de TI",
            font=FONT_TINY, bg=BG, fg=FG_MUTED
        ).pack(pady=(0, 18))

        # Usuário
        tk.Label(
            self, text="Usuário:", font=FONT, bg=BG, fg=FG_DIM
        ).pack(pady=(0, 3))
        self.nome_entry = tk.Entry(
            self, font=(FONT_FAMILY, 12), bg=BG_ENTRY, fg=FG,
            insertbackground=FG, relief="flat", bd=0,
            highlightthickness=2, highlightbackground=BORDER,
            highlightcolor=ACCENT, justify="center", width=28
        )
        self.nome_entry.pack(ipady=7, pady=(0, 10))
        self.nome_entry.focus_set()

        # Senha
        tk.Label(
            self, text="Senha:", font=FONT, bg=BG, fg=FG_DIM
        ).pack(pady=(0, 3))
        self.senha_entry = tk.Entry(
            self, font=(FONT_FAMILY, 12), bg=BG_ENTRY, fg=FG,
            insertbackground=FG, relief="flat", bd=0,
            highlightthickness=2, highlightbackground=BORDER,
            highlightcolor=ACCENT, justify="center", width=28,
            show="●"
        )
        self.senha_entry.pack(ipady=7, pady=(0, 16))
        self.senha_entry.bind("<Return>", lambda e: self._entrar())

        btn = tk.Button(
            self, text="Entrar", font=FONT_BOLD,
            bg=ACCENT, fg=HEADER_BG, activebackground=ACCENT2,
            activeforeground=HEADER_BG, relief="flat", bd=0,
            cursor="hand2", command=self._entrar, padx=40
        )
        btn.pack(ipady=6)

        # Créditos discretos
        tk.Label(
            self, text="Developed by: Wallison, Jonas",
            font=(FONT_FAMILY, 8), bg=BG, fg=FG_MUTED
        ).pack(side="bottom", pady=(0, 8))

    def _entrar(self):
        nome = self.nome_entry.get().strip()
        senha = self.senha_entry.get().strip()
        if not nome or not senha:
            messagebox.showwarning("Atenção", "Informe usuário e senha.")
            return
        user = database.autenticar_usuario(nome, senha)
        if not user:
            messagebox.showerror("Erro", "Usuário ou senha incorretos.")
            self.senha_entry.delete(0, "end")
            self.senha_entry.focus_set()
            return
        self.usuario = user["usuario"]
        self.is_super = user["is_super"] == 1
        self.is_viewer = (user["usuario"] == "Consultor")
        self.destroy()


class App(tk.Tk):
    def __init__(self, usuario: str, is_super: bool = False,
                 is_viewer: bool = False):
        super().__init__()
        self.title("SESAT — Controle de Equipamentos")
        self.geometry("1500x860")
        self.state("zoomed")  # Tela cheia (maximizado)
        self.configure(bg=BG)
        self.minsize(1100, 650)

        self.editing_id = None
        self._api_data = {}
        self.usuario = usuario
        self.is_super = is_super
        self.is_viewer = is_viewer
        self._logout_requested = False
        self._opcoes_ordenacao = {
            "Mais recentes": "mais_recentes",
            "Mais antigos": "mais_antigos",
            "Saídas mais recentes": "saida_mais_recentes",
            "Saídas mais antigas": "saida_mais_antigos",
            "Tombamento A-Z": "tombamento_az",
            "Tombamento Z-A": "tombamento_za",
        }

        self._setup_styles()
        self._build_ui()
        self._carregar_tabela()

        # Log de login
        database.registrar_log(self.usuario, "LOGIN",
                               detalhes="Usuário entrou no sistema")

        # Registrar logout ao fechar a janela pelo X
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # Fallback: atexit garante log de logout mesmo se destroy falhar
        self._logged_out = False
        atexit.register(self._atexit_logout)

        # Travar interface para viewer (somente leitura)
        if self.is_viewer:
            self._aplicar_modo_viewer()

    # ═══════════════════════════════════════════════════════════════
    #  Estilos ttk
    # ═══════════════════════════════════════════════════════════════
    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use("clam")

        # Notebook (abas)
        s.configure("TNotebook", background=BG, borderwidth=0)
        s.configure("TNotebook.Tab",
                    background=BG_SURFACE, foreground=FG_DIM,
                    font=FONT_BOLD, padding=(18, 8), borderwidth=0)
        s.map("TNotebook.Tab",
              background=[("selected", BG_CARD), ("active", BG_HOVER)],
              foreground=[("selected", ACCENT), ("active", FG)])

        # Treeview — bordas visíveis em cada célula
        CELL_BORDER = "#6272a4"
        s.configure("Custom.Treeview",
                    background=BG_CARD, foreground=FG, fieldbackground=BG_CARD,
                    font=FONT_SMALL, rowheight=30,
                    borderwidth=1, relief="solid",
                    bordercolor=CELL_BORDER,
                    lightcolor=CELL_BORDER, darkcolor=CELL_BORDER)
        s.configure("Custom.Treeview.Heading",
                    background=HEADER_BG, foreground=ACCENT,
                    font=FONT_BOLD, borderwidth=1, relief="solid",
                    padding=(6, 4))
        s.map("Custom.Treeview",
              background=[("selected", SEL_BG)],
              foreground=[("selected", FG)])
        s.map("Custom.Treeview.Heading",
              background=[("active", BG_HOVER)])

        # Layout padrão mantido para que o tema clam renderize as divisórias entre células
        s.layout("Custom.Treeview", [
            ("Custom.Treeview.border", {"sticky": "nswe", "children": [
                ("Custom.Treeview.padding", {"sticky": "nswe", "children": [
                    ("Custom.Treeview.treearea", {"sticky": "nswe"})
                ]})
            ]})
        ])

        # Scrollbar
        s.configure("Vertical.TScrollbar",
                    background=BG_SURFACE, troughcolor=BG_CARD,
                    arrowcolor=FG_DIM, borderwidth=0)
        s.configure("Horizontal.TScrollbar",
                    background=BG_SURFACE, troughcolor=BG_CARD,
                    arrowcolor=FG_DIM, borderwidth=0)

        # Combobox
        s.configure("TCombobox",
                    fieldbackground=BG_ENTRY, background=BG_ENTRY,
                    foreground=FG, selectbackground=SEL_BG,
                    selectforeground=FG, borderwidth=0)
        s.map("TCombobox",
              fieldbackground=[("readonly", BG_ENTRY)],
              foreground=[("readonly", FG)])

    # ═══════════════════════════════════════════════════════════════
    #  Construção da UI
    # ═══════════════════════════════════════════════════════════════
    def _build_ui(self):
        # ── Barra de título ──
        self._build_header()

        # ── Container principal com grid ──
        container = tk.Frame(self, bg=BG)
        container.pack(fill="both", expand=True, padx=12, pady=(0, 0))
        container.columnconfigure(0, weight=0, minsize=430)
        container.columnconfigure(1, weight=1, minsize=500)
        container.rowconfigure(0, weight=1)

        # ── FORMULÁRIO (esquerda) ──
        self._build_form_panel(container)

        # ── TABELA (direita) ──
        self._build_table_panel(container)

        # ── Status bar ──
        self._build_status_bar()

    def _build_header(self):
        hdr = tk.Frame(self, bg=HEADER_BG, height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        inner = tk.Frame(hdr, bg=HEADER_BG)
        inner.pack(fill="both", expand=True, padx=20)

        tk.Label(
            inner, text="⚙", font=(FONT_FAMILY, 22), bg=HEADER_BG, fg=ACCENT
        ).pack(side="left", padx=(0, 8))

        title_block = tk.Frame(inner, bg=HEADER_BG)
        title_block.pack(side="left")
        tk.Label(
            title_block, text="SESAT", font=FONT_TITLE, bg=HEADER_BG, fg=FG
        ).pack(anchor="w")
        tk.Label(
            title_block, text="Sistema de Entrada e Saída de Ativos de TI",
            font=FONT_TINY, bg=HEADER_BG, fg=FG_MUTED
        ).pack(anchor="w")

        # Usuário logado + logout
        user_frame = tk.Frame(inner, bg=HEADER_BG)
        user_frame.pack(side="right", padx=(0, 5))

        # Relógio
        self._clock_var = tk.StringVar()
        tk.Label(
            user_frame, textvariable=self._clock_var, font=FONT_SMALL,
            bg=HEADER_BG, fg=FG_DIM
        ).pack(anchor="e")

        user_row = tk.Frame(user_frame, bg=HEADER_BG)
        user_row.pack(anchor="e")

        tk.Label(
            user_row, text=f"👤 {self.usuario}",
            font=FONT_SMALL, bg=HEADER_BG, fg=ACCENT2
        ).pack(side="left", padx=(0, 8))

        btn_logout = tk.Button(
            user_row, text="🚪 Sair", font=FONT_TINY,
            bg=HEADER_BG, fg=RED, activebackground=BG_HOVER,
            activeforeground=RED, relief="flat", bd=0,
            cursor="hand2", command=self._on_logout, padx=6
        )
        btn_logout.pack(side="left")
        btn_logout.bind("<Enter>", lambda e: btn_logout.config(bg=BG_HOVER))
        btn_logout.bind("<Leave>", lambda e: btn_logout.config(bg=HEADER_BG))

        # Botão gerenciar usuários (só para Supervisor)
        if self.is_super:
            btn_users = tk.Button(
                user_row, text="👥 Usuários", font=FONT_TINY,
                bg=HEADER_BG, fg=YELLOW, activebackground=BG_HOVER,
                activeforeground=YELLOW, relief="flat", bd=0,
                cursor="hand2", command=self._on_gerenciar_usuarios, padx=6
            )
            btn_users.pack(side="left", padx=(8, 0))
            btn_users.bind("<Enter>", lambda e: btn_users.config(bg=BG_HOVER))
            btn_users.bind("<Leave>", lambda e: btn_users.config(bg=HEADER_BG))

        self._update_clock()

    def _update_clock(self):
        self._clock_var.set(datetime.now().strftime("%d/%m/%Y  %H:%M:%S"))
        self.after(1000, self._update_clock)

    # ───────────────────────────────────────────────────────────────
    #  Painel do Formulário (esquerda) — com scroll correto
    # ───────────────────────────────────────────────────────────────
    def _build_form_panel(self, parent):
        outer = tk.Frame(parent, bg=BORDER)
        outer.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=0)
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        # Canvas para scroll
        self._form_canvas = tk.Canvas(outer, bg=BG_CARD, highlightthickness=0,
                                      width=420)
        scrollbar = ttk.Scrollbar(outer, orient="vertical",
                                  command=self._form_canvas.yview)

        # Frame interno que contém todo o formulário
        self._form_inner = tk.Frame(self._form_canvas, bg=BG_CARD)

        # Janela dentro do canvas
        self._form_canvas_win = self._form_canvas.create_window(
            (0, 0), window=self._form_inner, anchor="nw")

        # Quando o conteúdo muda de tamanho, atualiza a scrollregion
        self._form_inner.bind("<Configure>", self._on_form_configure)
        # Quando o canvas muda de tamanho, ajusta a largura do frame interno
        self._form_canvas.bind("<Configure>", self._on_canvas_configure)

        self._form_canvas.configure(yscrollcommand=scrollbar.set)

        # Layout com grid para controle preciso
        self._form_canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Scroll com roda do mouse — binds locais (não globais)
        self._form_canvas.bind("<Enter>", self._bind_form_scroll)
        self._form_canvas.bind("<Leave>", self._unbind_form_scroll)
        self._form_inner.bind("<Enter>", self._bind_form_scroll)

        # Construir o conteúdo do formulário
        self._build_form_contents(self._form_inner)

    def _on_form_configure(self, event):
        self._form_canvas.configure(scrollregion=self._form_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # Faz o frame interno ter a mesma largura do canvas visível
        self._form_canvas.itemconfig(self._form_canvas_win, width=event.width)

    def _bind_form_scroll(self, event):
        self._form_canvas.bind_all(
            "<Button-4>",
            lambda e: self._form_canvas.yview_scroll(-3, "units"))
        self._form_canvas.bind_all(
            "<Button-5>",
            lambda e: self._form_canvas.yview_scroll(3, "units"))
        self._form_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self._form_canvas.yview_scroll(
                int(-1 * (e.delta / 120)), "units"))

    def _unbind_form_scroll(self, event):
        self._form_canvas.unbind_all("<Button-4>")
        self._form_canvas.unbind_all("<Button-5>")
        self._form_canvas.unbind_all("<MouseWheel>")

    # ───────────────────────────────────────────────────────────────
    #  Conteúdo do Formulário
    # ───────────────────────────────────────────────────────────────
    def _build_form_contents(self, f):
        self.entries = {}

        # ── Seção: Dados do Chamado (com botão Limpar) ──
        self._section_label(f, "  REGISTRO DE EQUIPAMENTO",
                            ACCENT, with_clear=True)

        # Tombamento + botão consultar lado a lado
        tk.Label(f, text="Nº Tombamento *", **LABEL_OPTS).pack(
            fill="x", padx=PADX, pady=PADY_FIELD)

        tomb_row = tk.Frame(f, bg=BG_CARD)
        tomb_row.pack(fill="x", padx=PADX, pady=(0, 4))
        tomb_row.columnconfigure(0, weight=1)

        tomb_entry = tk.Entry(tomb_row, **ENTRY_OPTS)
        tomb_entry.grid(row=0, column=0, sticky="ew", ipady=5)
        self.entries["tombamento"] = tomb_entry

        self.btn_consultar = tk.Button(
            tomb_row, text="🔍 Consultar", font=FONT_BOLD,
            bg=ACCENT, fg=HEADER_BG, activebackground=ACCENT2,
            activeforeground=HEADER_BG, relief="flat", bd=0,
            cursor="hand2", command=self._on_consultar, padx=12
        )
        self.btn_consultar.grid(
            row=0, column=1, sticky="ns", padx=(8, 0), ipady=3)

        # ── Data de Entrada ──
        tk.Label(f, text="Data de Entrada", **LABEL_OPTS).pack(
            fill="x", padx=PADX, pady=PADY_FIELD)

        dt_frame = tk.Frame(f, bg=BG_CARD)
        dt_frame.pack(fill="x", padx=PADX, pady=(0, 4))

        self._data_entrada_mode = tk.StringVar(value="auto")

        rb_auto = tk.Radiobutton(
            dt_frame, text="Automática (data do sistema)",
            variable=self._data_entrada_mode, value="auto",
            font=FONT_SMALL, bg=BG_CARD, fg=FG_DIM,
            selectcolor=BG_ENTRY, activebackground=BG_CARD,
            activeforeground=FG, indicatoron=True, anchor="w",
            command=self._toggle_data_entrada
        )
        rb_auto.pack(fill="x", pady=(0, 1))

        rb_manual = tk.Radiobutton(
            dt_frame, text="Manual (informar data)",
            variable=self._data_entrada_mode, value="manual",
            font=FONT_SMALL, bg=BG_CARD, fg=FG_DIM,
            selectcolor=BG_ENTRY, activebackground=BG_CARD,
            activeforeground=FG, indicatoron=True, anchor="w",
            command=self._toggle_data_entrada
        )
        rb_manual.pack(fill="x", pady=(0, 2))

        # Frame que agrupa entry + botão calendário (começa oculto)
        self._data_entrada_frame = tk.Frame(f, bg=BG_CARD)
        # NÃO pack agora — _toggle_data_entrada controla visibilidade

        self._data_entrada_entry = tk.Entry(
            self._data_entrada_frame, **ENTRY_OPTS)
        self._data_entrada_entry.pack(
            side="left", fill="x", expand=True, ipady=5)
        self.entries["data_entrada"] = self._data_entrada_entry

        btn_cal = tk.Button(
            self._data_entrada_frame, text="📅", font=FONT_BOLD,
            bg=BG_ENTRY, fg=ACCENT, activebackground=BG_HOVER,
            activeforeground=ACCENT2, relief="flat", bd=0,
            cursor="hand2",
            command=lambda: self._abrir_calendario(self._data_entrada_entry),
            width=3
        )
        btn_cal.pack(side="left", padx=(4, 0), ipady=5)

        # Âncora invisível para posicionar o frame de data corretamente
        self._data_entrada_anchor = tk.Frame(f, height=0, bg=BG_CARD)
        self._data_entrada_anchor.pack(fill="x")

        # Campos editáveis simples
        simple_fields = [
            ("num_chamado",           "Nº Chamado"),
            ("tecnico_opr_entrada",   "Técnico OPR Entrada"),
            ("tecnico_sesat",         "Técnico SESAT"),
            ("tecnico_opr_devolucao", "Técnico OPR Devolução"),
        ]
        for key, label in simple_fields:
            tk.Label(f, text=label, **LABEL_OPTS).pack(
                fill="x", padx=PADX, pady=PADY_FIELD)
            entry = tk.Entry(f, **ENTRY_OPTS)
            entry.pack(fill="x", padx=PADX, pady=(0, 4), ipady=5)
            self.entries[key] = entry

        # ── Data de Saída (com calendário) ──
        tk.Label(f, text="Data de Saída (DD/MM/AA)", **LABEL_OPTS).pack(
            fill="x", padx=PADX, pady=PADY_FIELD)
        self._data_saida_frame = tk.Frame(f, bg=BG_CARD)
        self._data_saida_frame.pack(fill="x", padx=PADX, pady=(0, 4))

        self._data_saida_entry = tk.Entry(
            self._data_saida_frame, **ENTRY_OPTS)
        self._data_saida_entry.pack(
            side="left", fill="x", expand=True, ipady=5)
        self.entries["data_saida"] = self._data_saida_entry

        btn_cal_saida = tk.Button(
            self._data_saida_frame, text="📅", font=FONT_BOLD,
            bg=BG_ENTRY, fg=ACCENT, activebackground=BG_HOVER,
            activeforeground=ACCENT2, relief="flat", bd=0,
            cursor="hand2",
            command=lambda: self._abrir_calendario(self._data_saida_entry),
            width=3
        )
        btn_cal_saida.pack(side="left", padx=(4, 0), ipady=5)

        # Campos de texto (multilinha)
        for key, label in [("observacoes", "Observações"),
                           ("anotacoes", "Anotações Diversas")]:
            tk.Label(f, text=label, **LABEL_OPTS).pack(
                fill="x", padx=PADX, pady=PADY_FIELD)
            txt = tk.Text(
                f, height=2, font=FONT, bg=BG_ENTRY, fg=FG,
                insertbackground=FG, relief="flat", bd=0,
                highlightthickness=2, highlightbackground="#6272a4",
                highlightcolor=ACCENT, wrap="word", padx=6, pady=4
            )
            txt.pack(fill="x", padx=PADX, pady=(0, 4))
            self.entries[key] = txt

        # ── Separador ──
        self._divider(f)

        # ── Seção: Dados do Patrimônio (API ou manual) ──
        self._section_label(f, "  DADOS DO PATRIMÔNIO (API / Manual)", YELLOW)

        # Container para o aviso + campos de patrimônio
        api_section = tk.Frame(f, bg=BG_CARD)
        api_section.pack(fill="x")

        # Aviso de modo manual (oculto por padrão)
        self._manual_warning = tk.Label(
            api_section,
            text="⚠ Patrimônio não encontrado na API.\n"
                 "Preencha os campos abaixo manualmente.",
            font=FONT_SMALL, bg="#4a3a20", fg=YELLOW, anchor="w",
            padx=10, pady=5, wraplength=390, justify="left"
        )
        # Não pack agora — será mostrado quando necessário

        self.info_entries = {}
        info_fields = [
            ("secao_zona",         "Seção / Zona"),
            ("nome_setor",         "Nome do Setor"),
            ("local_setor",        "Endereço"),
            ("tipo",               "Tipo / Desc. do Bem"),
            ("descricao_completa", "Descrição Completa"),
            ("nome_responsavel",   "Responsável"),
        ]
        for key, label in info_fields:
            tk.Label(api_section, text=label, **LABEL_OPTS).pack(
                fill="x", padx=PADX, pady=(5, 0))
            entry = tk.Entry(
                api_section, font=FONT_SMALL, bg=BG_ENTRY, fg=GREEN,
                insertbackground=GREEN, relief="flat", bd=0,
                highlightthickness=2, highlightbackground="#6272a4",
                highlightcolor=YELLOW, disabledbackground=BG_ENTRY,
                disabledforeground=FG_MUTED, state="disabled"
            )
            entry.pack(fill="x", padx=PADX, pady=(2, 2), ipady=5)
            self.info_entries[key] = entry

        # ── Separador ──
        self._divider(f)

        # ── Botões de ação ──
        btn_frame = tk.Frame(f, bg=BG_CARD)
        btn_frame.pack(fill="x", padx=PADX, pady=(8, 18))

        self.btn_salvar = self._action_button(
            btn_frame, "💾  Salvar Registro", GREEN, HEADER_BG, self._on_salvar)
        self.btn_salvar.pack(fill="x", pady=(0, 5), ipady=7)

        self.btn_excluir = self._action_button(
            btn_frame, "🗑️ Excluir", RED, None, self._on_excluir)
        self.btn_excluir.pack(fill="x", ipady=5)

    # ── Widgets auxiliares do form ──

    def _section_label(self, parent, text, color, with_clear=False):
        frame = tk.Frame(parent, bg=BG_SURFACE)
        frame.pack(fill="x", padx=0, pady=(10, 6))
        tk.Label(
            frame, text=text, font=FONT_SUB, bg=BG_SURFACE, fg=color,
            anchor="w", padx=PADX, pady=6
        ).pack(side="left", fill="x", expand=True)
        if with_clear:
            action_frame = tk.Frame(frame, bg=BG_SURFACE)
            action_frame.pack(side="right", padx=(0, PADX))

            self.btn_atualizar = tk.Button(
                action_frame, text="🔄 Atualizar", font=FONT_SMALL,
                bg=BG_SURFACE, fg=ACCENT2, activebackground=BG_HOVER,
                activeforeground=FG, relief="flat", bd=0,
                cursor="hand2", command=self._on_atualizar_tabela, padx=10
            )
            self.btn_atualizar.pack(side="right", padx=(8, 0))
            self.btn_atualizar.bind("<Enter>",
                                    lambda e: self.btn_atualizar.config(bg=BG_HOVER))
            self.btn_atualizar.bind("<Leave>",
                                    lambda e: self.btn_atualizar.config(bg=BG_SURFACE))

            self.btn_limpar = tk.Button(
                action_frame, text="🧹 Novo / Limpar", font=FONT_SMALL,
                bg=BG_SURFACE, fg=FG_DIM, activebackground=BG_HOVER,
                activeforeground=FG, relief="flat", bd=0,
                cursor="hand2", command=self._on_limpar, padx=10
            )
            self.btn_limpar.pack(side="right")
            self.btn_limpar.bind("<Enter>",
                                 lambda e: self.btn_limpar.config(bg=BG_HOVER))
            self.btn_limpar.bind("<Leave>",
                                 lambda e: self.btn_limpar.config(bg=BG_SURFACE))

    def _divider(self, parent):
        tk.Frame(parent, bg=BORDER, height=1).pack(
            fill="x", padx=PADX, pady=(10, 2))

    def _action_button(self, parent, text, fg_color, bg_color, command):
        bg = bg_color or BG_ENTRY
        btn = tk.Button(
            parent, text=text, font=FONT_BOLD,
            bg=bg, fg=fg_color, activebackground=BG_HOVER,
            activeforeground=fg_color, relief="flat", bd=0,
            cursor="hand2", command=command
        )
        original_bg = bg
        btn.bind("<Enter>", lambda e: btn.config(bg=BG_HOVER))
        btn.bind("<Leave>", lambda e: btn.config(bg=original_bg))
        return btn

    # ───────────────────────────────────────────────────────────────
    #  Painel da Tabela (direita)
    # ───────────────────────────────────────────────────────────────
    def _build_table_panel(self, parent):
        outer = tk.Frame(parent, bg=BG_CARD, highlightthickness=1,
                         highlightbackground=BORDER)
        outer.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=0)
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)

        self._build_search_bar(outer)
        self._build_table(outer)

    def _build_search_bar(self, parent):
        bar = tk.Frame(parent, bg=BG_SURFACE)
        bar.grid(row=0, column=0, sticky="ew")

        inner = tk.Frame(bar, bg=BG_SURFACE)
        inner.pack(fill="x", padx=12, pady=8)

        tk.Label(inner, text="🔎", font=(FONT_FAMILY, 12),
                 bg=BG_SURFACE, fg=FG_DIM).pack(side="left")

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._carregar_tabela())
        tk.Entry(
            inner, textvariable=self.search_var, font=FONT,
            bg=BG_ENTRY, fg=FG, insertbackground=FG, relief="flat",
            highlightthickness=2, highlightbackground=BORDER,
            highlightcolor=ACCENT, width=28
        ).pack(side="left", padx=(6, 12), ipady=4)

        tk.Label(inner, text="Filtrar por:", font=FONT_SMALL,
                 bg=BG_SURFACE, fg=FG_DIM).pack(side="left", padx=(0, 5))

        self.search_field = ttk.Combobox(inner, values=[
            "todos", "tombamento", "zona", "tipo", "chamado",
            "opr_entrada", "sesat", "opr_devolucao", "observacoes"
        ], state="readonly", width=14, font=FONT_SMALL)
        self.search_field.set("todos")
        self.search_field.pack(side="left", padx=(0, 10))
        self.search_field.bind("<<ComboboxSelected>>",
                               lambda *_: self._carregar_tabela())

        tk.Label(inner, text="Ordenar:", font=FONT_SMALL,
                 bg=BG_SURFACE, fg=FG_DIM).pack(side="left", padx=(0, 5))

        self.sort_field = ttk.Combobox(
            inner,
            values=list(self._opcoes_ordenacao.keys()),
            state="readonly",
            width=20,
            font=FONT_SMALL,
        )
        self.sort_field.set("Mais recentes")
        self.sort_field.pack(side="left", padx=(0, 10))
        self.sort_field.bind("<<ComboboxSelected>>",
                             lambda *_: self._carregar_tabela())

        if not self.is_viewer:
            btn_export = tk.Button(
                inner, text="📊 Exportar XLSX", font=FONT_BOLD,
                bg=BG_ENTRY, fg=YELLOW, activebackground=BG_HOVER,
                activeforeground=YELLOW, relief="flat", bd=0,
                cursor="hand2", command=self._on_exportar, padx=14
            )
            btn_export.pack(side="right", ipady=3)
            btn_export.bind(
                "<Enter>", lambda e: btn_export.config(bg=BG_HOVER))
            btn_export.bind(
                "<Leave>", lambda e: btn_export.config(bg=BG_ENTRY))

        if self.is_super:
            btn_import = tk.Button(
                inner, text="📥 Importar XLSX", font=FONT_BOLD,
                bg=BG_ENTRY, fg=GREEN, activebackground=BG_HOVER,
                activeforeground=GREEN, relief="flat", bd=0,
                cursor="hand2", command=self._on_importar, padx=14
            )
            btn_import.pack(side="right", padx=(0, 6), ipady=3)
            btn_import.bind(
                "<Enter>", lambda e: btn_import.config(bg=BG_HOVER))
            btn_import.bind(
                "<Leave>", lambda e: btn_import.config(bg=BG_ENTRY))

            btn_export_log = tk.Button(
                inner, text="📋 Exportar Log", font=FONT_BOLD,
                bg=BG_ENTRY, fg=PEACH, activebackground=BG_HOVER,
                activeforeground=PEACH, relief="flat", bd=0,
                cursor="hand2", command=self._on_exportar_log, padx=14
            )
            btn_export_log.pack(side="right", padx=(0, 6), ipady=3)
            btn_export_log.bind(
                "<Enter>", lambda e: btn_export_log.config(bg=BG_HOVER))
            btn_export_log.bind(
                "<Leave>", lambda e: btn_export_log.config(bg=BG_ENTRY))

        btn_ver_log = tk.Button(
            inner, text="📜 Ver Logs", font=FONT_BOLD,
            bg=BG_ENTRY, fg=MAUVE, activebackground=BG_HOVER,
            activeforeground=MAUVE, relief="flat", bd=0,
            cursor="hand2", command=self._on_ver_logs, padx=14
        )
        btn_ver_log.pack(side="right", padx=(0, 6), ipady=3)
        btn_ver_log.bind("<Enter>", lambda e: btn_ver_log.config(bg=BG_HOVER))
        btn_ver_log.bind("<Leave>", lambda e: btn_ver_log.config(bg=BG_ENTRY))

    def _build_table(self, parent):
        cols = (
            "id", "data_entrada", "secao_zona", "tombamento", "tipo",
            "num_chamado", "tecnico_opr_entrada", "tecnico_sesat",
            "tecnico_opr_devolucao", "data_saida", "observacoes", "anotacoes"
        )
        col_names = (
            "ID", "Entrada", "Seção/Zona", "Tomb.", "Tipo",
            "Nº Chamado", "OPR Entrada", "SESAT",
            "OPR Devol.", "Saída", "Observações", "Anotações"
        )
        col_widths = (0, 82, 110, 80, 155, 115, 115, 100, 115, 82, 135, 165)

        tree_frame = tk.Frame(parent, bg=BG, highlightthickness=0, bd=0)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=8, pady=(4, 8))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="browse", style="Custom.Treeview"
        )
        for c, name, w in zip(cols, col_names, col_widths):
            self.tree.heading(c, text=name, anchor="center")
            if c == "id":
                # Coluna oculta — usada internamente
                self.tree.column(c, width=0, minwidth=0, stretch=False)
            else:
                self.tree.column(c, width=w, minwidth=35, stretch=True,
                                 anchor="center")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.tree.tag_configure("even", background=ROW_EVEN)
        self.tree.tag_configure("odd", background=ROW_ODD)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", self._on_select)

    # ───────────────────────────────────────────────────────────────
    #  Status bar
    # ───────────────────────────────────────────────────────────────
    def _build_status_bar(self):
        self.status_var = tk.StringVar(value="Pronto.")
        bar = tk.Frame(self, bg=HEADER_BG, height=28)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self.status_label = tk.Label(
            bar, textvariable=self.status_var, font=FONT_TINY,
            bg=HEADER_BG, fg=FG_MUTED, anchor="w", padx=20
        )
        self.status_label.pack(fill="both", expand=True)

    # ═══════════════════════════════════════════════════════════════
    #  Helpers
    # ═══════════════════════════════════════════════════════════════
    def _set_status(self, msg, color=FG_DIM):
        self.status_var.set(msg)
        self.status_label.config(fg=color)

    def _get_entry_value(self, key):
        widget = self.entries[key]
        if isinstance(widget, tk.Text):
            return widget.get("1.0", "end-1c").strip()
        return widget.get().strip()

    def _set_entry_value(self, key, value):
        widget = self.entries[key]
        if isinstance(widget, tk.Text):
            widget.delete("1.0", "end")
            widget.insert("1.0", value or "")
        else:
            widget.delete(0, "end")
            widget.insert(0, value or "")

    # ═══════════════════════════════════════════════════════════════
    #  Modo Viewer (somente leitura)
    # ═══════════════════════════════════════════════════════════════
    def _aplicar_modo_viewer(self):
        """Desabilita todos os campos e botões de edição para o Viewer."""
        # Desabilitar todos os campos de entrada do formulário
        for key, widget in self.entries.items():
            if isinstance(widget, tk.Text):
                widget.config(state="disabled")
            else:
                widget.config(state="disabled",
                              disabledforeground=FG_MUTED)

        # Desabilitar campos de informação da API
        for entry in self.info_entries.values():
            entry.config(state="disabled", disabledforeground=FG_MUTED)

        # Desabilitar botões de ação do formulário
        self.btn_salvar.config(state="disabled", fg=FG_MUTED)
        self.btn_excluir.config(state="disabled", fg=FG_MUTED)
        self.btn_consultar.config(state="disabled", fg=FG_MUTED)
        if hasattr(self, "btn_limpar"):
            self.btn_limpar.config(state="disabled", fg=FG_MUTED)

        # Desabilitar radio buttons da data de entrada
        for w in self._data_entrada_frame.master.winfo_children():
            if isinstance(w, tk.Radiobutton):
                w.config(state="disabled")

        self._set_status("👁️ Modo visualização — somente leitura", YELLOW)

    # ═══════════════════════════════════════════════════════════════
    #  Ações do formulário
    # ═══════════════════════════════════════════════════════════════
    def _toggle_data_entrada(self):
        """Mostra/esconde o frame de data de entrada conforme o modo."""
        if self._data_entrada_mode.get() == "manual":
            self._data_entrada_frame.pack(
                fill="x", padx=PADX, pady=(0, 4),
                before=self._data_entrada_anchor)
            self._data_entrada_entry.config(state="normal")
            self._data_entrada_entry.delete(0, "end")
            self._data_entrada_entry.focus_set()
        else:
            self._data_entrada_entry.config(state="normal")
            self._data_entrada_entry.delete(0, "end")
            self._data_entrada_frame.pack_forget()

    # ── Calendário pop-up ──
    def _abrir_calendario(self, target_entry=None):
        """Abre um pop-up de calendário mensal para seleção de data."""
        if target_entry is not None:
            self._cal_target_entry = target_entry

        # Se já existe um calendário aberto, fechar
        if hasattr(self, '_cal_win') and self._cal_win is not None:
            try:
                self._cal_win.destroy()
            except tk.TclError:
                pass
            self._cal_win = None
            return

        hoje = datetime.now()
        self._cal_ano = hoje.year
        self._cal_mes = hoje.month

        self._cal_win = tk.Toplevel(self)
        self._cal_win.title("Selecionar Data")
        self._cal_win.configure(bg=BG)
        self._cal_win.resizable(False, False)
        self._cal_win.overrideredirect(True)

        # Posicionar próximo ao campo
        x = self._cal_target_entry.winfo_rootx()
        y = self._cal_target_entry.winfo_rooty() + \
            self._cal_target_entry.winfo_height() + 4
        self._cal_win.geometry(f"+{x}+{y}")

        # Borda
        border_frame = tk.Frame(self._cal_win, bg=BORDER, bd=1, relief="solid")
        border_frame.pack(fill="both", expand=True)

        self._cal_container = tk.Frame(border_frame, bg=BG)
        self._cal_container.pack(fill="both", expand=True, padx=1, pady=1)

        self._desenhar_calendario()

        # Fechar ao clicar em qualquer lugar fora do calendário
        self._cal_win.update_idletasks()
        self.bind_all("<Button-1>", self._fechar_calendario_click, add="+")
        self._cal_win.protocol("WM_DELETE_WINDOW", self._fechar_calendario)

    def _desenhar_calendario(self):
        """Redesenha o conteúdo do calendário para o mês/ano atual."""
        for w in self._cal_container.winfo_children():
            w.destroy()

        MESES = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio",
                 "Junho", "Julho", "Agosto", "Setembro", "Outubro",
                 "Novembro", "Dezembro"]

        # ── Navegação mês ──
        nav = tk.Frame(self._cal_container, bg=BG)
        nav.pack(fill="x", pady=(6, 4), padx=6)

        tk.Button(
            nav, text="◀", font=FONT_BOLD, bg=BG, fg=ACCENT,
            activebackground=BG_HOVER, activeforeground=ACCENT2,
            relief="flat", bd=0, cursor="hand2",
            command=self._cal_mes_anterior
        ).pack(side="left")

        tk.Label(
            nav, text=f"{MESES[self._cal_mes]} {self._cal_ano}",
            font=FONT_BOLD, bg=BG, fg=FG
        ).pack(side="left", expand=True)

        tk.Button(
            nav, text="▶", font=FONT_BOLD, bg=BG, fg=ACCENT,
            activebackground=BG_HOVER, activeforeground=ACCENT2,
            relief="flat", bd=0, cursor="hand2",
            command=self._cal_mes_proximo
        ).pack(side="right")

        # ── Cabeçalho dias da semana ──
        grid = tk.Frame(self._cal_container, bg=BG)
        grid.pack(padx=6, pady=(0, 6))

        dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        for c, ds in enumerate(dias_semana):
            tk.Label(
                grid, text=ds, font=FONT_TINY, bg=BG, fg=FG_MUTED,
                width=4
            ).grid(row=0, column=c, padx=1, pady=(0, 2))

        # ── Dias do mês ──
        hoje = datetime.now()
        matriz = cal_mod.monthcalendar(self._cal_ano, self._cal_mes)

        for r, semana in enumerate(matriz, 1):
            for c, dia in enumerate(semana):
                if dia == 0:
                    tk.Label(
                        grid, text="", font=FONT_SMALL, bg=BG,
                        width=4
                    ).grid(row=r, column=c, padx=1, pady=1)
                else:
                    is_hoje = (dia == hoje.day and
                               self._cal_mes == hoje.month and
                               self._cal_ano == hoje.year)
                    fg_c = HEADER_BG if is_hoje else FG
                    bg_c = ACCENT if is_hoje else BG_ENTRY

                    btn = tk.Button(
                        grid, text=str(dia), font=FONT_SMALL,
                        bg=bg_c, fg=fg_c,
                        activebackground=ACCENT2,
                        activeforeground=HEADER_BG,
                        relief="flat", bd=0, width=4, cursor="hand2",
                        command=lambda d=dia: self._selecionar_dia(d)
                    )
                    btn.grid(row=r, column=c, padx=1, pady=1)

    def _cal_mes_anterior(self):
        self._cal_mes -= 1
        if self._cal_mes < 1:
            self._cal_mes = 12
            self._cal_ano -= 1
        self._desenhar_calendario()

    def _cal_mes_proximo(self):
        self._cal_mes += 1
        if self._cal_mes > 12:
            self._cal_mes = 1
            self._cal_ano += 1
        self._desenhar_calendario()

    def _selecionar_dia(self, dia):
        data_str = f"{dia:02d}/{self._cal_mes:02d}/{self._cal_ano % 100:02d}"
        self._cal_target_entry.delete(0, "end")
        self._cal_target_entry.insert(0, data_str)
        self._fechar_calendario()

    def _fechar_calendario(self):
        """Fecha o calendário e remove o binding global."""
        try:
            self.unbind_all("<Button-1>")
        except tk.TclError:
            pass
        try:
            if self._cal_win and self._cal_win.winfo_exists():
                self._cal_win.destroy()
        except tk.TclError:
            pass
        self._cal_win = None

    def _fechar_calendario_click(self, event):
        """Fecha o calendário se o clique foi fora dele."""
        if self._cal_win is None:
            return
        try:
            w = event.widget
            # Se o clique é no próprio calendário ou em filhos dele, ignorar
            if w == self._cal_win or str(w).startswith(str(self._cal_win)):
                return
            self._fechar_calendario()
        except tk.TclError:
            self._fechar_calendario()

    def _on_consultar(self):
        tomb = self._get_entry_value("tombamento")
        if not tomb:
            messagebox.showwarning(
                "Atenção", "Informe o número do tombamento.")
            return

        self._set_status("⏳ Consultando patrimônio na intranet...", YELLOW)
        self.btn_consultar.config(state="disabled", text="⏳ ...")

        def _fetch():
            dados = consulta_api.consultar_patrimonio(tomb)
            self.after(0, lambda: self._preencher_api(dados))

        threading.Thread(target=_fetch, daemon=True).start()

    def _preencher_api(self, dados):
        self.btn_consultar.config(state="normal", text="🔍 Consultar")
        if dados is None:
            # Modo manual: habilitar campos para preenchimento
            self._set_status(
                "⚠ Patrimônio não encontrado na API — preencha os dados manualmente.",
                YELLOW)
            # Mostrar aviso no topo da seção
            self._manual_warning.pack_forget()
            siblings = self._manual_warning.master.pack_slaves()
            if siblings:
                self._manual_warning.pack(fill="x", padx=PADX, pady=(4, 6),
                                          before=siblings[0])
            else:
                self._manual_warning.pack(fill="x", padx=PADX, pady=(4, 6))

            for entry in self.info_entries.values():
                entry.config(state="normal", fg=YELLOW,
                             highlightbackground=YELLOW)
            self._api_data = {}
            return

        # Sucesso: preencher e travar como somente leitura
        self._manual_warning.pack_forget()

        mapping = {
            "secao_zona":         dados.get("sigla", ""),
            "nome_setor":         dados.get("nome_setor", ""),
            "local_setor":        dados.get("local", ""),
            "tipo":               dados.get("ds_bem", ""),
            "descricao_completa": dados.get("descricao_completa", ""),
            "nome_responsavel":   dados.get("nome_responsavel", ""),
        }
        self._api_data = mapping

        for key, val in mapping.items():
            entry = self.info_entries.get(key)
            if entry:
                entry.config(state="normal")
                entry.delete(0, "end")
                entry.insert(0, val or "")
                entry.config(state="disabled", disabledforeground=GREEN,
                             highlightbackground=BORDER)

        self._set_status(
            f"✅ Patrimônio {dados.get('patrimonio', '')} encontrado com sucesso.",
            GREEN)

    def _on_salvar(self):
        if self.is_viewer:
            messagebox.showinfo(
                "Info", "Modo visualização — não é possível salvar.")
            return
        tomb = self._get_entry_value("tombamento")
        if not tomb:
            messagebox.showwarning(
                "Atenção", "Informe o número do tombamento.")
            return

        api = self._api_data or {}

        # Se os campos da API estão editáveis (modo manual), lê diretamente
        def _get_info(key):
            if api.get(key):
                return api[key]
            entry = self.info_entries.get(key)
            if entry:
                try:
                    return entry.get().strip()
                except tk.TclError:
                    return ""
            return ""

        # Determinar data de entrada
        if self._data_entrada_mode.get() == "manual":
            data_entrada_val = self._get_entry_value("data_entrada")
            if not data_entrada_val or data_entrada_val == "DD/MM/AA":
                messagebox.showwarning("Atenção",
                                       "Informe a data de entrada manualmente.")
                return
        else:
            data_entrada_val = datetime.now().strftime("%d/%m/%y")

        dados = {
            "data_entrada":         data_entrada_val,
            "secao_zona":           _get_info("secao_zona"),
            "nome_setor":           _get_info("nome_setor"),
            "local_setor":          _get_info("local_setor"),
            "tombamento":           tomb,
            "tipo":                 _get_info("tipo"),
            "descricao_completa":   _get_info("descricao_completa"),
            "nome_responsavel":     _get_info("nome_responsavel"),
            "valor_unitario":       "",
            "num_chamado":          self._get_entry_value("num_chamado"),
            "tecnico_opr_entrada":  self._get_entry_value("tecnico_opr_entrada"),
            "tecnico_sesat":        self._get_entry_value("tecnico_sesat"),
            "tecnico_opr_devolucao": self._get_entry_value("tecnico_opr_devolucao"),
            "data_saida":           self._get_entry_value("data_saida"),
            "observacoes":          self._get_entry_value("observacoes"),
            "anotacoes":            self._get_entry_value("anotacoes"),
        }

        try:
            if self.editing_id:
                original = database.buscar_por_id(self.editing_id)
                if original and self._data_entrada_mode.get() == "auto":
                    dados["data_entrada"] = original["data_entrada"]
                database.atualizar_equipamento(self.editing_id, dados)
                msg = f"✅ Tombamento {tomb} atualizado com sucesso."
                database.registrar_log(
                    self.usuario, "ATUALIZAÇÃO",
                    tombamento=tomb,
                    data_entrada=dados["data_entrada"],
                    data_saida=dados.get("data_saida", ""),
                    detalhes=f"Registro atualizado pelo usuário"
                )
            else:
                new_id = database.inserir_equipamento(dados)
                msg = f"✅ Tombamento {tomb} salvo com sucesso."
                database.registrar_log(
                    self.usuario, "CADASTRO",
                    tombamento=tomb,
                    data_entrada=dados["data_entrada"],
                    data_saida=dados.get("data_saida", ""),
                    detalhes=f"Novo registro cadastrado"
                )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar:\n{e}")
            return

        # Limpar formulário e desselecionar tabela
        self._on_limpar()
        self._carregar_tabela()
        # Desselecionar qualquer linha na tabela
        for sel in self.tree.selection():
            self.tree.selection_remove(sel)
        # Restaurar mensagem de sucesso (após o limpar)
        self._set_status(msg, GREEN)

    def _on_limpar(self):
        for key, widget in self.entries.items():
            if key == "data_entrada":
                continue  # tratado separadamente abaixo
            if isinstance(widget, tk.Text):
                widget.delete("1.0", "end")
            else:
                widget.delete(0, "end")
        # Limpar campos de patrimônio e travar
        for entry in self.info_entries.values():
            entry.config(state="normal")
            entry.delete(0, "end")
            entry.config(state="disabled", disabledforeground=FG_MUTED,
                         highlightbackground=BORDER)
        self._manual_warning.pack_forget()
        self._api_data = {}
        self.editing_id = None
        # Resetar data de entrada para automático
        self._data_entrada_mode.set("auto")
        self._toggle_data_entrada()
        self.btn_salvar.config(text="💾  Salvar Registro")
        self._set_status("Formulário limpo.", FG_DIM)

    def _on_atualizar_tabela(self):
        """Recarrega a tabela manualmente para refletir alterações recentes."""
        self._carregar_tabela()
        self._set_status("🔄 Tabela atualizada.", ACCENT)

    def _on_excluir(self):
        if self.is_viewer:
            messagebox.showinfo(
                "Info", "Modo visualização — não é possível excluir.")
            return
        if not self.editing_id:
            messagebox.showinfo("Info",
                                "Selecione um registro na tabela para excluir.")
            return
        tomb = self._get_entry_value("tombamento") or self.editing_id

        # Pop-up para selecionar motivo da exclusão
        motivo = self._pedir_motivo_exclusao(tomb)
        if motivo is None:
            return  # Cancelou

        database.registrar_log(
            self.usuario, "EXCLUSÃO",
            tombamento=str(tomb),
            detalhes=f"Motivo: {motivo}"
        )
        database.deletar_equipamento(self.editing_id)
        self._on_limpar()
        self._carregar_tabela()
        for sel in self.tree.selection():
            self.tree.selection_remove(sel)
        self._set_status(f"🗑️ Tombamento {tomb} excluído.", RED)

    def _pedir_motivo_exclusao(self, tomb):
        """Abre janela para o usuário escolher o motivo da exclusão."""
        win = tk.Toplevel(self)
        win.title("Motivo da Exclusão")
        win.geometry("450x360")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # Centralizar
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 225
        y = self.winfo_y() + (self.winfo_height() // 2) - 180
        win.geometry(f"450x360+{x}+{y}")

        resultado = {"motivo": None}

        tk.Label(
            win, text="🗑️ Excluir Registro", font=FONT_SUB,
            bg=BG, fg=RED
        ).pack(pady=(16, 4))
        tk.Label(
            win, text=f"Tombamento: {tomb}", font=FONT,
            bg=BG, fg=FG_DIM
        ).pack(pady=(0, 12))
        tk.Label(
            win, text="Selecione o motivo da exclusão:",
            font=FONT, bg=BG, fg=FG
        ).pack(pady=(0, 10))

        motivo_var = tk.StringVar(value="")

        motivos = [
            "Chamado finalizado",
            "Cadastrado de forma errada",
            "Chamado cancelado",
            "Outro",
        ]

        for m in motivos:
            rb = tk.Radiobutton(
                win, text=m, variable=motivo_var, value=m,
                font=FONT, bg=BG, fg=FG, selectcolor=BG_ENTRY,
                activebackground=BG, activeforeground=FG,
                indicatoron=True, anchor="w", padx=20,
                tristatevalue="_none_",
                command=lambda: _toggle_outro()
            )
            rb.pack(fill="x", padx=40, pady=2)

        # Campo de texto para "Outro"
        outro_frame = tk.Frame(win, bg=BG)
        outro_frame.pack(fill="x", padx=60, pady=(4, 0))
        outro_entry = tk.Entry(
            outro_frame, font=FONT, bg=BG_ENTRY, fg=FG,
            insertbackground=FG, relief="flat", bd=0,
            highlightthickness=2, highlightbackground=BORDER,
            highlightcolor=ACCENT, state="disabled"
        )
        outro_entry.pack(fill="x", ipady=5)

        def _toggle_outro():
            if motivo_var.get() == "Outro":
                outro_entry.config(state="normal")
                outro_entry.focus_set()
            else:
                outro_entry.delete(0, "end")
                outro_entry.config(state="disabled")

        def _confirmar():
            sel = motivo_var.get()
            if not sel:
                messagebox.showwarning("Atenção", "Selecione um motivo.",
                                       parent=win)
                return
            if sel == "Outro":
                texto = outro_entry.get().strip()
                if not texto:
                    messagebox.showwarning("Atenção",
                                           "Descreva o motivo da exclusão.",
                                           parent=win)
                    return
                resultado["motivo"] = f"Outro: {texto}"
            else:
                resultado["motivo"] = sel
            win.destroy()

        def _cancelar():
            win.destroy()

        btn_frame = tk.Frame(win, bg=BG)
        btn_frame.pack(pady=(14, 12))

        tk.Button(
            btn_frame, text="✅ Confirmar Exclusão", font=FONT_BOLD,
            bg=BG_ENTRY, fg=RED, activebackground=BG_HOVER,
            activeforeground=RED, relief="flat", bd=0,
            cursor="hand2", command=_confirmar, padx=16
        ).pack(side="left", padx=6, ipady=5)

        tk.Button(
            btn_frame, text="Cancelar", font=FONT_BOLD,
            bg=BG_ENTRY, fg=FG_DIM, activebackground=BG_HOVER,
            activeforeground=FG, relief="flat", bd=0,
            cursor="hand2", command=_cancelar, padx=16
        ).pack(side="left", padx=6, ipady=5)

        win.wait_window()
        return resultado["motivo"]

    # ═══════════════════════════════════════════════════════════════
    #  Seleção na tabela → preencher formulário
    # ═══════════════════════════════════════════════════════════════
    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        vals = item["values"]
        if not vals:
            return

        equip_id = vals[0]
        reg = database.buscar_por_id(equip_id)
        if not reg:
            return

        self.editing_id = equip_id
        tomb_val = reg.get("tombamento", equip_id)
        self.btn_salvar.config(text=f"💾  Atualizar {tomb_val}")

        field_map = {
            "tombamento": "tombamento",
            "num_chamado": "num_chamado",
            "tecnico_opr_entrada": "tecnico_opr_entrada",
            "tecnico_sesat": "tecnico_sesat",
            "tecnico_opr_devolucao": "tecnico_opr_devolucao",
            "data_saida": "data_saida",
            "observacoes": "observacoes",
            "anotacoes": "anotacoes",
        }
        for form_key, db_key in field_map.items():
            self._set_entry_value(form_key, reg.get(db_key, ""))

        # Preencher data de entrada no modo manual
        data_ent = reg.get("data_entrada", "")
        if data_ent:
            self._data_entrada_mode.set("manual")
            self._toggle_data_entrada()
            self._data_entrada_entry.delete(0, "end")
            self._data_entrada_entry.insert(0, data_ent)

        api_fields = {
            "secao_zona":         reg.get("secao_zona", ""),
            "nome_setor":         reg.get("nome_setor", ""),
            "local_setor":        reg.get("local_setor", ""),
            "tipo":               reg.get("tipo", ""),
            "descricao_completa": reg.get("descricao_completa", ""),
            "nome_responsavel":   reg.get("nome_responsavel", ""),
        }
        self._api_data = api_fields

        # Ocultar aviso manual
        self._manual_warning.pack_forget()

        for key, val in api_fields.items():
            entry = self.info_entries.get(key)
            if entry:
                entry.config(state="normal")
                entry.delete(0, "end")
                entry.insert(0, val or "")
                entry.config(state="disabled", disabledforeground=GREEN if val else FG_MUTED,
                             highlightbackground=BORDER)

        self._set_status(f"Editando tombamento {tomb_val}", ACCENT)

    # ═══════════════════════════════════════════════════════════════
    #  Carregar dados na tabela
    # ═══════════════════════════════════════════════════════════════
    def _carregar_tabela(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        filtro = self.search_var.get() if hasattr(self, "search_var") else ""
        campo = self.search_field.get() if hasattr(self, "search_field") else "todos"
        ordenacao_label = (
            self.sort_field.get()
            if hasattr(self, "sort_field") and self.sort_field.get()
            else "Mais recentes"
        )
        ordenacao = self._opcoes_ordenacao.get(
            ordenacao_label, "mais_recentes")
        registros = database.buscar_todos(filtro, campo, ordenacao)

        for i, reg in enumerate(registros):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=(
                reg["id"],
                reg.get("data_entrada", ""),
                reg.get("secao_zona", ""),
                reg.get("tombamento", ""),
                reg.get("tipo", ""),
                reg.get("num_chamado", ""),
                reg.get("tecnico_opr_entrada", ""),
                reg.get("tecnico_sesat", ""),
                reg.get("tecnico_opr_devolucao", ""),
                reg.get("data_saida", ""),
                reg.get("observacoes", ""),
                reg.get("anotacoes", ""),
            ), tags=(tag,))

        total = len(registros)
        self._set_status(
            f"{total} registro(s) encontrado(s) · ordem: {ordenacao_label}.",
            FG_DIM,
        )

    # ═══════════════════════════════════════════════════════════════
    #  Importar planilha (apenas Supervisor)
    # ═══════════════════════════════════════════════════════════════
    def _on_importar(self):
        if not self.is_super:
            messagebox.showerror("Acesso negado",
                                 "Apenas o Supervisor pode importar planilhas.")
            return

        caminho = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            title="Selecionar planilha para importar"
        )
        if not caminho:
            return

        if not messagebox.askyesno(
                "Confirmar Importação",
                f"Deseja importar os dados da planilha?\n\n"
                f"{caminho}\n\n"
                "Registros duplicados (mesmo tombamento + data de entrada) "
                "serão ignorados automaticamente."):
            return

        self._set_status("⏳ Importando planilha... aguarde.", YELLOW)
        self.update_idletasks()

        def _importar_thread():
            try:
                resultado = importar.importar_xlsx(caminho, self.usuario)
                self.after(0, lambda: self._importar_concluido(resultado))
            except Exception as e:
                self.after(0, lambda: self._importar_erro(str(e)))

        threading.Thread(target=_importar_thread, daemon=True).start()

    def _importar_concluido(self, resultado):
        self._carregar_tabela()

        msg = (
            f"Importação concluída!\n\n"
            f"Abas processadas: {resultado['total_abas']}\n"
            f"Registros importados: {resultado['total_importados']}\n"
            f"Duplicados ignorados: {resultado['total_duplicados']}\n"
            f"Linhas vazias ignoradas: {resultado['total_ignorados']}"
        )
        if resultado["erros"]:
            msg += f"\n\nErros ({len(resultado['erros'])}):\n"
            msg += "\n".join(resultado["erros"][:10])
            if len(resultado["erros"]) > 10:
                msg += f"\n... e mais {len(resultado['erros']) - 10} erro(s)"

        self._set_status(
            f"📥 Importação: {resultado['total_importados']} registros importados.",
            GREEN)
        messagebox.showinfo("Importação Concluída", msg)

    def _importar_erro(self, erro):
        self._set_status("❌ Falha na importação.", RED)
        messagebox.showerror("Erro na Importação",
                             f"Falha ao importar planilha:\n{erro}")

    # ═══════════════════════════════════════════════════════════════
    #  Exportar planilha
    # ═══════════════════════════════════════════════════════════════
    def _on_exportar(self):
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"PLANILHA_OPR_SESAT_{datetime.now().strftime('%Y%m%d')}.xlsx",
            title="Salvar planilha como..."
        )
        if not caminho:
            return
        try:
            exportar.exportar_xlsx(caminho)
            self._set_status(f"📊 Planilha exportada: {caminho}", GREEN)
            messagebox.showinfo("Sucesso",
                                f"Planilha exportada com sucesso!\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar:\n{e}")

    # ═══════════════════════════════════════════════════════════════
    #  Ver Logs
    # ═══════════════════════════════════════════════════════════════
    def _on_ver_logs(self):
        win = tk.Toplevel(self)
        win.title("SESAT — Logs de Atividade")
        win.geometry("950x550")
        win.configure(bg=BG)
        win.minsize(700, 400)
        win.transient(self)

        # Barra de busca
        top = tk.Frame(win, bg=BG_SURFACE)
        top.pack(fill="x")
        inner = tk.Frame(top, bg=BG_SURFACE)
        inner.pack(fill="x", padx=12, pady=8)

        tk.Label(inner, text="📜 Logs de Atividade", font=FONT_SUB,
                 bg=BG_SURFACE, fg=MAUVE).pack(side="left")

        search_var = tk.StringVar()
        tk.Entry(
            inner, textvariable=search_var, font=FONT,
            bg=BG_ENTRY, fg=FG, insertbackground=FG, relief="flat",
            highlightthickness=2, highlightbackground=BORDER,
            highlightcolor=ACCENT, width=25
        ).pack(side="right", padx=(6, 0), ipady=4)
        tk.Label(inner, text="🔎", font=(FONT_FAMILY, 11),
                 bg=BG_SURFACE, fg=FG_DIM).pack(side="right")

        # Treeview de logs
        cols = ("data_hora", "usuario", "acao", "tombamento",
                "data_entrada", "data_saida", "detalhes")
        col_names = ("Data/Hora", "Usuário", "Ação", "Tombamento",
                     "Dt. Entrada", "Dt. Saída", "Detalhes")
        col_widths = (140, 110, 100, 100, 90, 90, 260)

        tree_frame = tk.Frame(win, bg=BG_CARD)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        log_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="browse", style="Custom.Treeview"
        )
        for c, name, w in zip(cols, col_names, col_widths):
            log_tree.heading(c, text=name)
            log_tree.column(c, width=w, minwidth=40, stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=log_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                            command=log_tree.xview)
        log_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        log_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        log_tree.tag_configure("even", background=ROW_EVEN)
        log_tree.tag_configure("odd", background=ROW_ODD)

        def _carregar_logs(*_):
            for item in log_tree.get_children():
                log_tree.delete(item)
            logs = database.buscar_logs(search_var.get())
            for i, log in enumerate(logs):
                tag = "even" if i % 2 == 0 else "odd"
                log_tree.insert("", "end", values=(
                    log.get("data_hora", ""),
                    log.get("usuario", ""),
                    log.get("acao", ""),
                    log.get("tombamento", ""),
                    log.get("data_entrada", ""),
                    log.get("data_saida", ""),
                    log.get("detalhes", ""),
                ), tags=(tag,))

        search_var.trace_add("write", _carregar_logs)
        _carregar_logs()

        # Status
        status = tk.Label(win, text="", font=FONT_TINY,
                          bg=HEADER_BG, fg=FG_MUTED, anchor="w", padx=20)
        status.pack(fill="x", side="bottom")
        total = len(database.buscar_logs())
        status.config(
            text=f"{total} registro(s) de log (mantidos por 6 meses)")

    # ═══════════════════════════════════════════════════════════════
    #  Exportar Log
    # ═══════════════════════════════════════════════════════════════
    def _on_exportar_log(self):
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"LOG_SESAT_{datetime.now().strftime('%Y%m%d')}.xlsx",
            title="Salvar log como..."
        )
        if not caminho:
            return
        try:
            self._exportar_log_xlsx(caminho)
            self._set_status(f"📋 Log exportado: {caminho}", GREEN)
            messagebox.showinfo("Sucesso",
                                f"Log exportado com sucesso!\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar log:\n{e}")

    def _exportar_log_xlsx(self, caminho: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        logs = database.buscar_logs()
        wb = Workbook()
        ws = wb.active
        ws.title = "LOG SESAT"

        headers = ["DATA/HORA", "USUÁRIO", "AÇÃO", "TOMBAMENTO",
                   "DATA ENTRADA", "DATA SAÍDA", "DETALHES"]

        header_font = Font(name="Liberation Sans",
                           bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(
            start_color="6C3483", end_color="6C3483", fill_type="solid")
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        campo_map = ["data_hora", "usuario", "acao", "tombamento",
                     "data_entrada", "data_saida", "detalhes"]

        data_font = Font(name="Liberation Sans", size=10)
        data_alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, log in enumerate(logs, 2):
            for col_idx, campo in enumerate(campo_map, 1):
                valor = log.get(campo, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=valor or "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        col_widths = [20, 18, 14, 14, 14, 14, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        wb.save(caminho)

    # ═══════════════════════════════════════════════════════════════
    #  Gerenciar Usuários (apenas Supervisor)
    # ═══════════════════════════════════════════════════════════════
    def _on_gerenciar_usuarios(self):
        if not self.is_super:
            return

        win = tk.Toplevel(self)
        win.title("SESAT — Gerenciar Usuários")
        win.geometry("600x500")
        win.configure(bg=BG)
        win.minsize(500, 400)
        win.transient(self)

        # Título
        top = tk.Frame(win, bg=BG_SURFACE)
        top.pack(fill="x")
        tk.Label(top, text="👥 Gerenciar Usuários", font=FONT_SUB,
                 bg=BG_SURFACE, fg=YELLOW, padx=12, pady=8).pack(side="left")

        # Formulário de adição
        form = tk.Frame(win, bg=BG_CARD)
        form.pack(fill="x", padx=12, pady=(12, 6))

        tk.Label(form, text="Novo usuário:", font=FONT_SMALL,
                 bg=BG_CARD, fg=FG_DIM).grid(row=0, column=0, padx=(8, 4), pady=6, sticky="e")
        new_user_entry = tk.Entry(form, font=FONT, bg=BG_ENTRY, fg=FG,
                                  insertbackground=FG, relief="flat",
                                  highlightthickness=2, highlightbackground=BORDER,
                                  highlightcolor=ACCENT, width=18)
        new_user_entry.grid(row=0, column=1, padx=4, pady=6, ipady=4)

        tk.Label(form, text="Senha:", font=FONT_SMALL,
                 bg=BG_CARD, fg=FG_DIM).grid(row=0, column=2, padx=(8, 4), pady=6, sticky="e")
        new_pass_entry = tk.Entry(form, font=FONT, bg=BG_ENTRY, fg=FG,
                                  insertbackground=FG, relief="flat",
                                  highlightthickness=2, highlightbackground=BORDER,
                                  highlightcolor=ACCENT, width=18, show="●")
        new_pass_entry.grid(row=0, column=3, padx=4, pady=6, ipady=4)

        # Treeview de usuários
        cols = ("id", "usuario", "tipo", "criado_em")
        col_names = ("ID", "Usuário", "Tipo", "Criado em")
        col_widths = (0, 180, 100, 160)

        tree_frame = tk.Frame(win, bg=BG_CARD)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        user_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="browse", style="Custom.Treeview"
        )
        for c, name, w in zip(cols, col_names, col_widths):
            user_tree.heading(c, text=name)
            if c == "id":
                user_tree.column(c, width=0, minwidth=0, stretch=False)
            else:
                user_tree.column(c, width=w, minwidth=40, stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=user_tree.yview)
        user_tree.configure(yscrollcommand=vsb.set)
        user_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        user_tree.tag_configure("even", background=ROW_EVEN)
        user_tree.tag_configure("odd", background=ROW_ODD)
        user_tree.tag_configure("super", background="#3a2a4a")

        def _carregar_usuarios():
            for item in user_tree.get_children():
                user_tree.delete(item)
            usuarios = database.listar_usuarios()
            for i, u in enumerate(usuarios):
                tipo = "Supervisor" if u["is_super"] else "Usuário"
                tag = "super" if u["is_super"] else (
                    "even" if i % 2 == 0 else "odd")
                user_tree.insert("", "end", values=(
                    u["id"], u["usuario"], tipo, u.get("created_at", "")
                ), tags=(tag,))

        def _adicionar():
            nome = new_user_entry.get().strip()
            senha = new_pass_entry.get().strip()
            if not nome or not senha:
                messagebox.showwarning("Atenção", "Informe usuário e senha.",
                                       parent=win)
                return
            if len(senha) < 4:
                messagebox.showwarning("Atenção", "A senha deve ter pelo menos 4 caracteres.",
                                       parent=win)
                return
            ok = database.criar_usuario(nome, senha)
            if ok:
                messagebox.showinfo("Sucesso", f"Usuário '{nome}' criado!",
                                    parent=win)
                new_user_entry.delete(0, "end")
                new_pass_entry.delete(0, "end")
                _carregar_usuarios()
            else:
                messagebox.showerror("Erro",
                                     f"Usuário '{nome}' já existe.",
                                     parent=win)

        def _excluir():
            sel = user_tree.selection()
            if not sel:
                messagebox.showinfo("Info", "Selecione um usuário para excluir.",
                                    parent=win)
                return
            vals = user_tree.item(sel[0])["values"]
            uid = vals[0]
            nome = vals[1]
            if not messagebox.askyesno("Confirmar",
                                       f"Excluir o usuário '{nome}'?",
                                       parent=win):
                return
            ok = database.excluir_usuario(uid)
            if ok:
                _carregar_usuarios()
            else:
                messagebox.showerror("Erro",
                                     "Não é possível excluir o Supervisor.",
                                     parent=win)

        def _resetar_senha():
            sel = user_tree.selection()
            if not sel:
                messagebox.showinfo("Info", "Selecione um usuário.",
                                    parent=win)
                return
            vals = user_tree.item(sel[0])["values"]
            uid = vals[0]
            nome = vals[1]
            nova = new_pass_entry.get().strip()
            if not nova:
                messagebox.showwarning("Atenção",
                                       "Digite a nova senha no campo 'Senha'.",
                                       parent=win)
                return
            if len(nova) < 4:
                messagebox.showwarning("Atenção",
                                       "A senha deve ter pelo menos 4 caracteres.",
                                       parent=win)
                return
            database.alterar_senha_usuario(uid, nova)
            messagebox.showinfo("Sucesso",
                                f"Senha de '{nome}' alterada!",
                                parent=win)
            new_pass_entry.delete(0, "end")

        # Botões
        btn_frame = tk.Frame(form, bg=BG_CARD)
        btn_frame.grid(row=1, column=0, columnspan=4, pady=(6, 8))

        btn_add = tk.Button(
            btn_frame, text="➕ Adicionar", font=FONT_BOLD,
            bg=BG_ENTRY, fg=GREEN, activebackground=BG_HOVER,
            activeforeground=GREEN, relief="flat", bd=0,
            cursor="hand2", command=_adicionar, padx=12
        )
        btn_add.pack(side="left", padx=4, ipady=4)

        btn_del = tk.Button(
            btn_frame, text="🗑️ Excluir", font=FONT_BOLD,
            bg=BG_ENTRY, fg=RED, activebackground=BG_HOVER,
            activeforeground=RED, relief="flat", bd=0,
            cursor="hand2", command=_excluir, padx=12
        )
        btn_del.pack(side="left", padx=4, ipady=4)

        btn_reset = tk.Button(
            btn_frame, text="🔑 Resetar Senha", font=FONT_BOLD,
            bg=BG_ENTRY, fg=PEACH, activebackground=BG_HOVER,
            activeforeground=PEACH, relief="flat", bd=0,
            cursor="hand2", command=_resetar_senha, padx=12
        )
        btn_reset.pack(side="left", padx=4, ipady=4)

        _carregar_usuarios()

    # ═══════════════════════════════════════════════════════════════
    #  Logout / Fechar
    # ═══════════════════════════════════════════════════════════════
    def _on_logout(self):
        if not messagebox.askyesno(
                "Logout", f"Deseja sair como {self.usuario}?"):
            return
        self._do_logout("Usuário saiu do sistema")
        self._logout_requested = True
        self.destroy()

    def _on_close(self):
        """Chamado ao fechar a janela pelo botão X — pede confirmação."""
        if not messagebox.askyesno(
                "Fechar aplicação",
                f"Deseja encerrar o SESAT?\n\n"
                f"O usuário '{self.usuario}' será desconectado."):
            return
        self._do_logout("Usuário saiu do sistema (fechou o programa)")

        self.destroy()

    def _do_logout(self, detalhes: str):
        """Registra logout no banco se ainda não foi registrado."""
        if self._logged_out:
            return
        self._logged_out = True
        try:
            database.registrar_log(
                self.usuario, "LOGOUT", detalhes=detalhes
            )
        except Exception:
            pass  # Evita erro ao fechar

    def _atexit_logout(self):
        """Fallback: garante log de logout ao encerrar o processo."""
        self._do_logout("Usuário saiu do sistema (processo encerrado)")


# ═══════════════════════════════════════════════════════════════════
#  Execução — loop de login/logout
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    while True:
        login = LoginWindow()
        login.mainloop()

        if not login.usuario:
            break  # Fechou a janela de login sem entrar

        app = App(login.usuario, login.is_super, login.is_viewer)
        app.mainloop()

        if not app._logout_requested:
            break  # Fechou o app normalmente (sem logout)
