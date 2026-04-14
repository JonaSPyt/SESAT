"""
Módulo de exportação dos dados para planilha Excel (.xlsx)
compatível com o formato da planilha OPR - SESAT existente.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database


def exportar_xlsx(caminho: str, registros: list | None = None):
    """
    Exporta os registros para um arquivo .xlsx no formato da planilha original.
    Se registros não for fornecido, exporta todos do banco.
    """
    if registros is None:
        registros = database.buscar_todos()

    wb = Workbook()
    ws = wb.active
    ws.title = "PLANILHA OPR - SESAT"

    # ── Cabeçalhos (mesma ordem da planilha original) ──
    headers = [
        "DATA DA ENTRADA",
        "SEÇÃO/ZONA",
        "TOMBAMENTO",
        "TIPO",
        "Nº CHAMADO",
        "TÉCNICO OPR ENTRADA",
        "TÉCNICO SESAT",
        "TÉCNICO OPR DEVOLUÇÃO",
        "DATA DA SAÍDA",
        "OBSERVAÇÕES",
        "ANOTAÇÕES DIVERSAS",
    ]

    # Estilo do cabeçalho
    header_font = Font(name="Liberation Sans", bold=True, size=10, color="000000")
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin")

    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Mapeamento dos campos do banco para as colunas
    campo_map = [
        "data_entrada",
        "secao_zona",
        "tombamento",
        "tipo",
        "num_chamado",
        "tecnico_opr_entrada",
        "tecnico_sesat",
        "tecnico_opr_devolucao",
        "data_saida",
        "observacoes",
        "anotacoes",
    ]

    data_font = Font(name="Liberation Sans", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, campo in enumerate(campo_map, 1):
            valor = reg.get(campo, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor or "")
            cell.font = data_font
            cell.alignment = data_alignment

    # Bordas apenas internas na área de dados (sem moldura externa).
    first_data_row = 2
    last_data_row = len(registros) + 1
    last_col = len(campo_map)
    if last_data_row >= first_data_row:
        for row_idx in range(first_data_row, last_data_row + 1):
            for col_idx in range(1, last_col + 1):
                right = thin_side if col_idx < last_col else None
                bottom = thin_side if row_idx < last_data_row else None
                ws.cell(row=row_idx, column=col_idx).border = Border(
                    right=right,
                    bottom=bottom,
                )

    # Ajustar largura das colunas
    col_widths = [15, 18, 14, 22, 16, 20, 18, 22, 15, 25, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Congelar painel no cabeçalho
    ws.freeze_panes = "A2"

    wb.save(caminho)
    return caminho
